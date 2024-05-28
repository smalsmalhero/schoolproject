# 載入LineBot所需要的套件
from flask import Flask, request, abort, json
import requests

import json

# 多線程與排程
import threading
from datetime import datetime
import schedule
import time

# Excel
import openpyxl
from openpyxl import load_workbook

# LINE_bot
from linebot import (
    LineBotApi, WebhookHandler
)
from linebot.models import *

app = Flask(__name__)

access_token = 'ZgmUooVkJxW6AY570r0wJkTJNKrcV2/LU5ICIQ+R+I5fBgudLLpiLBjkzhH2ERtIoiMz02ObINOzrjMGvlX2jHRUnrCjF9fsaZu7aZFzZDZShsbIwbcJ4VxX7uVd9aIgQBCEe8QgkNq2rn6xD2B+ZAdB04t89/1O/w1cDnyilFU='
channel_secret = 'e9d2a7178b0eb60b3a30e4916a97e75f'

# 必須放上自己的Channel Access Token
line_bot_api = LineBotApi(access_token)
# 必須放上自己的Channel Secret
handler = WebhookHandler(channel_secret)

# LINE push 訊息函式
def push_message(msg, uid, token):
    headers = {'Authorization':f'Bearer {token}','Content-Type':'application/json'}   
    body = {
    'to':uid,
    'messages':[{
            "type": "text",
            "text": msg
        }]
    }
    req = requests.request('POST', 'https://api.line.me/v2/bot/message/push', headers=headers,data=json.dumps(body).encode('utf-8'))
    print(req.text)
# LINE reply 訊息函式
def reply_message(msg, rk, token):
    headers = {'Authorization':f'Bearer {token}','Content-Type':'application/json'}
    body = {
    'replyToken':rk,
    'messages':[{
            "type": "text",
            "text": msg
        }]
    }
    requests.request('POST', 'https://api.line.me/v2/bot/message/reply', headers=headers,data=json.dumps(body).encode('utf-8'))

# 設定UID
def set_UID(user_id, user_name):
    wb = openpyxl.load_workbook('data.xlsx', data_only=True)    # 打開檔案
    s1 = wb['take_madicine']                                   # 開啟take_madincine工作表

    user = 1
    while True:
        if s1[f'A{user+1}'].value == user_name:
            return '設置：已有相同的名稱'
        elif s1[f'A{user}'].value is None:
            s1[f'A{user}'].value = user_id
            s1[f'A{user+1}'].value = user_name
            break
        elif s1[f'A{user}'].value == user_id:
            s1[f'A{user+1}'].value = user_name
            break
        
        user += 4
    
    wb.save('data.xlsx')
# 讀取UID
def read_UID(user_name):
    wb = openpyxl.load_workbook('data.xlsx', data_only=True)  # 設定 data_only=True 只讀取計算後的數值

    s1 = wb['take_madicine']
    user = 2
    while True:
        read_data = s1[f'A{user}'].value
        read_data_2 = s1[f'A{user-1}'].value
        if read_data == user_name:
            user_id = s1[f'A{user-1}'].value
            break
        elif read_data_2 is None:
            user_id = ''
            break
        user +=4

    return user_id
# 設定定時器
def set_schelude(years, months, days, hours, minutes, user_id):
    def my_task():
        schedule.clear(user_id)
        print("任务执行时间：", time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        push_message('時間到該回診囉！', user_id, access_token)

    def run_scheduler():
        while True:
            schedule.run_pending()
            time.sleep(1)

    # 定义要执行任务的日期
    scheduled_date = datetime(int(years), int(months), int(days), int(hours), int(minutes))

    # 计算距离指定日期的时间差
    time_difference = scheduled_date - datetime.now()

    # 将时间差转换为秒数
    time_in_seconds = time_difference.total_seconds()

    # 定义一个定时任务，在指定日期执行
    schedule.every(time_in_seconds).seconds.do(my_task).tag(user_id)

    # 启动定时任务的线程
    scheduler_thread = threading.Thread(target=run_scheduler)
    scheduler_thread.start()
# 設定每日定時器
def set_daily(times, hour, minute, user_id):
    # 早上
    def my_morning_task():
        schedule.clear(f'{user_id}{times}')
        print("任务执行时间：", time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        if times == 1:
            push_message('早安 吃藥時間到囉! ฅ●ω●ฅ', user_id, access_token)
        elif times == 2:
           push_message('午安 吃藥時間到囉! ฅ●ω●ฅ', user_id, access_token)
        elif times == 3:
            push_message('晚安 吃藥時間到囉! ฅ●ω●ฅ', user_id, access_token)
        elif times == 4:
            push_message('睡覺前 也該吃藥囉! (⁠⁠ꈍ⁠ᴗ⁠ꈍ⁠)', user_id, access_token)

    # 執行排程的副程式
    def run_scheduler():
        while True:
            schedule.run_pending()
            time.sleep(1)

    schedule.clear(f'{user_id}{times}')

    if hour < 10 and minute < 10:
        schedule.every().days.at(f'0{hour}:0{minute}').do(my_morning_task).tag(f'{user_id}{times}')
    elif hour < 10:
        print(hour,minute)
        schedule.every().days.at(f'0{hour}:{minute}').do(my_morning_task).tag(f'{user_id}{times}')
    elif minute < 10:
        schedule.every().days.at(f'{hour}:0{minute}').do(my_morning_task).tag(f'{user_id}{times}')
    else:
        print(hour,minute)
        schedule.every().days.at(f'{hour}:{minute}').do(my_morning_task).tag(f'{user_id}{times}')

    # 啟動定時任務排程
    scheduler_thread = threading.Thread(target=run_scheduler)
    scheduler_thread.start()
# 設定吃藥時間
def set_take_medicine(user_id_data, user_name, time, hour, minute):
    wb = openpyxl.load_workbook('data.xlsx', data_only=True)    # 打開檔案
    s1 = wb['take_madicine']

    user_id_reflash = read_UID(user_name)
    user = 1
    print(user_id_reflash)
    msg = ''

    if user_id_reflash == '':
        while True:
            if s1[f'A{user}'].value is None:
                s1[f'A{user}'].value = user_id_data
                s1[f'B{time}'].value = hour
                s1[f'C{time}'].value = minute
                break
            elif s1[f'A{user}'].value == user_id_data:
                s1[f'B{user + time - 1}'].value = hour
                s1[f'C{user + time - 1}'].value = minute
                break
            user += 4
        
        wb.save('data.xlsx')

        send_take_medicine(user_id_data, time)
        msg = '已更新自己'
    else:
        while True:
            if s1[f'A{user+1}'].value == user_name:
                s1[f'B{user + time - 1}'].value = hour
                s1[f'C{user + time - 1}'].value = minute
                break
            user += 4
        wb.save('data.xlsx')

        send_take_medicine(user_id_reflash, time)
        msg = '已更新指定使用者'

    return msg
# 發送吃藥訊息
def send_take_medicine(user_id, time):
    wb = openpyxl.load_workbook('data.xlsx', data_only=True)  # 設定 data_only=True 只讀取計算後的數值

    s1 = wb['take_madicine']
    user = 1
    while True:
        if s1[f'A{user}'].value == user_id:
            hour = s1[f'B{user + time - 1}'].value
            minute = s1[f'C{user + time - 1}'].value
            set_daily(time, hour, minute, user_id)
            break
        user += 4
    

# 監聽所有來自 / 的 Post,GET Request
@app.route("/linebot",methods=['POST','GET'])
def linebot():
    body = request.get_data(as_text=True)
    try:        
        signature = request.headers['X-Line-Signature']             # 加入回傳的 headers
        handler.handle(body, signature)                             # 綁定訊息回傳的相關資訊
        json_data = json.loads(body)                                # 轉換內容為 json 格式
        reply_token = json_data['events'][0]['replyToken']          # 儲存 replyToken
        user_id = json_data['events'][0]['source']['userId']        # 儲存 user_id

        if 'message' in json_data['events'][0]:
            if json_data['events'][0]['message']['type'] == 'text':
                text = json_data['events'][0]['message']['text']
                text = text.replace('：',':').lstrip()
                print(text)
                if text[:7] == '設定:回診日期':
                    try:
                        date = text[7:].lstrip(' ')                                 # 單獨把日期移出來
                        year = int(date[:4].lstrip('0'))                            # 設定年
                        month = int(date[5:7].lstrip('0'))                          # 設定月
                        day = int(date[8:10].lstrip('0'))                           # 設定日
                        hour = int(date[11:13].lstrip('0'))                         # 設定時
                        if date[11:13] == '00':
                            hour = int(date[11:13].replace('0',' ',1).strip())        # 設定分
                        else:
                            hour = int(date[11:13].lstrip())
                        if date[14:16] == '00':
                            minute = int(date[14:16].replace('0',' ',1).strip())        # 設定分
                        else:
                            minute = int(date[14:16].lstrip())
                        set_schelude(year, month, day, hour, minute, user_id)
                        reply_message('設定：成功設置日期', reply_token, access_token)
                    except:
                        reply_message('錯誤訊息', reply_token, access_token)
                elif text[:7] == '設定:提醒時間':
                    try:
                        data = text[7:].lstrip(' ')
                        time = int(data[:2].lstrip('0'))
                        hour = int(data[3:5])       # 設定時
                        minute = int(data[6:8])     # 設定分
                        user = data[9:]             # 設定暱稱
                        print(time, hour, minute, user)

                        if hour >= 24 or hour < 0:
                            reply_message('時間設置錯誤', reply_token, access_token)
                        elif minute >= 60 or minute < 0:
                            reply_message('時間設置錯誤', reply_token, access_token)
                        
                        data_msg = set_take_medicine(user_id, user, time, hour, minute)
                        
                        word = ['早上','中午','晚上','睡前']
                        if hour < 10 and minute < 10:
                            reply_message(f'設定：成功設置\n' + data_msg + '\n' + f'{word[time-1]} 0{hour}點0{minute}分', reply_token, access_token)
                        elif hour < 10:
                            reply_message(f'設定：成功設置\n' + f'{data_msg}' + '\n' + f'{word[time-1]} 0{hour}點{minute}分', reply_token, access_token)
                        elif minute < 10:
                            reply_message(f'設定：成功設置\n' + f'{data_msg}' + '\n' + f'{word[time-1]} {hour}點0{minute}分', reply_token, access_token)
                        else:
                            reply_message(f'設定：成功設置\n' + f'{data_msg}' + '\n' + f'{word[time-1]} {hour}點{minute}分', reply_token, access_token)                                            
                    except Exception as e:
                        print('error', e)
                        reply_message('設定：錯誤訊息', reply_token, access_token)
                elif text[:8] == '設定:使用者名稱':
                    try:
                        data = text[8:].lstrip(' ')
                        if set_UID(user_id, data) != '設置：已有相同的名稱':
                            reply_message(f'設定：成功設置\n你的使用者名稱：\n{data}\n（可使用你的暱稱幫別人做設定）', reply_token, access_token)
                        else:
                            reply_message('設置：已有相同的名稱', reply_token, access_token)
                    except:
                        reply_message('設定錯誤', reply_token, access_token)     
        print('success')
    except Exception as e:
        print('error', e)
    return 'ok'
# 傳送資料給ESP32
@app.route('/data', methods=['POST','GET'])
def data():
    wb = openpyxl.load_workbook('data.xlsx', data_only=True)  # 設定 data_only=True 只讀取計算後的數值

    s1 = wb['take_madicine']

    # Python 的 dict 類型資料
    myDict = {
        "ID" : f'{s1["A1"].value}',
        "morning_time": [f'{s1["B1"].value}', f'{s1["C1"].value}'],
        "noon_time": [f'{s1["B2"].value}', f'{s1["C2"].value}'],
        "night_time": [f'{s1["B3"].value}', f'{s1["C3"].value}'],
        "beforebed_time": [f'{s1["B4"].value}', f'{s1["C4"].value}'],
    }
    myDict1 = {"user" : f'{myDict}'}
    myDict1 = json.dumps(myDict1, ensure_ascii=False)

    return myDict1
# 接收ESP32資料
@app.route('/<msg>')
def setup(msg):
    print(msg)
    push_message('已吃藥囉！', msg, access_token)
    return 'ok'

# 啟動應用程式
if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000)